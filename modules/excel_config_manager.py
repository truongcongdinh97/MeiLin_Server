#!/usr/bin/env python3
"""
Excel Configuration Manager for MeiLin
Export/Import user configurations via Excel files

Supports:
- Download template: Người dùng mới có thể tải về template
- Download config: Xuất cấu hình hiện tại ra Excel
- Upload config: Import cấu hình từ Excel
"""

import os
import io
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: pandas/openpyxl not installed. Excel features disabled.")

logger = logging.getLogger(__name__)


# ============================================================
# EXCEL TEMPLATE DEFINITIONS
# ============================================================

# Sheet definitions với cột và mô tả
TEMPLATE_SHEETS = {
    'Personality': {
        'description': 'Cấu hình tính cách của AI',
        'columns': [
            ('character_name', 'Tên nhân vật', 'MeiLin', 'Tên AI sẽ tự xưng'),
            ('wake_word', 'Wake Word', 'hi meilin', 'Từ khóa để gọi AI'),
            ('speaking_style', 'Phong cách nói', 'friendly', 'friendly/professional/cute/playful/formal'),
            ('primary_language', 'Ngôn ngữ', 'vi', 'vi/en/ja/zh/ko'),
            ('temperature', 'Temperature', '0.7', 'Độ sáng tạo (0.0-1.0)'),
            ('response_length', 'Độ dài trả lời', 'medium', 'short/medium/long'),
        ]
    },
    'Knowledge Base': {
        'description': 'Kiến thức cá nhân cho AI',
        'columns': [
            ('id', 'ID', 'KB_001', 'ID duy nhất'),
            ('category', 'Danh mục', 'Personal', 'Personal/Work/Hobby/etc'),
            ('priority', 'Độ ưu tiên', '1', '1=cao nhất, 5=thấp nhất'),
            ('content', 'Nội dung', 'Thông tin về bạn...', 'Kiến thức AI sẽ học'),
            ('tags', 'Tags', 'personal, info', 'Các tag phân loại'),
        ]
    },
    'Contacts': {
        'description': 'Danh bạ liên hệ cho voice commands',
        'columns': [
            ('uid', 'UID', 'U001', 'ID liên hệ'),
            ('name', 'Tên', 'Nguyễn Văn A', 'Tên hiển thị'),
            ('platform', 'Nền tảng', 'Zalo', 'Zalo/Telegram/Email/SMS'),
            ('phone', 'Số điện thoại', '0912345678', 'Số điện thoại'),
            ('email', 'Email', 'a@example.com', 'Địa chỉ email'),
            ('notes', 'Ghi chú', 'Bạn thân', 'Ghi chú thêm'),
        ]
    },
    'Workflows': {
        'description': 'Cấu hình tự động hóa',
        'columns': [
            ('workflow_id', 'ID Workflow', 'WF_001', 'ID duy nhất'),
            ('name', 'Tên workflow', 'Gửi tin nhắn', 'Tên hiển thị'),
            ('trigger', 'Trigger', 'voice_command', 'voice_command/schedule/event'),
            ('platform', 'Nền tảng', 'Zalo', 'Zalo/Telegram/Email/N8N'),
            ('action', 'Hành động', 'send_message', 'send_message/create_task/etc'),
            ('parameters', 'Tham số', 'uid,content', 'Các tham số cần'),
            ('enabled', 'Bật/Tắt', 'TRUE', 'TRUE/FALSE'),
        ]
    },
    'API Keys': {
        'description': 'API keys (được mã hóa khi import)',
        'columns': [
            ('provider', 'Provider', 'deepseek', 'deepseek/openai/anthropic/etc'),
            ('type', 'Loại', 'llm', 'llm/tts/stt'),
            ('api_key', 'API Key', 'sk-xxx...', 'API key của bạn'),
            ('api_base', 'API Base', '', 'URL tùy chỉnh (optional)'),
            ('model', 'Model', 'deepseek-chat', 'Tên model'),
            ('is_default', 'Mặc định', 'TRUE', 'TRUE/FALSE'),
        ]
    },
    'Instructions': {
        'description': 'Hướng dẫn sử dụng',
        'content': '''
📋 HƯỚNG DẪN SỬ DỤNG MEILIN CONFIGURATION TEMPLATE

1️⃣ PERSONALITY (Tính cách AI)
   - Đặt tên, wake word, phong cách nói cho AI
   - Chỉ điền 1 dòng dữ liệu

2️⃣ KNOWLEDGE BASE (Kiến thức)
   - Thêm thông tin cá nhân để AI hiểu bạn hơn
   - Có thể thêm nhiều dòng
   - Ví dụ: sở thích, công việc, gia đình...

3️⃣ CONTACTS (Danh bạ)
   - Danh sách liên hệ cho voice commands
   - "Gửi tin nhắn cho Anh Định" → tìm trong contacts

4️⃣ WORKFLOWS (Tự động hóa)
   - Cấu hình các hành động tự động
   - Kết nối với N8N, Zalo, Telegram...

5️⃣ API KEYS (Khóa API)
   - Thêm API keys cho LLM, TTS, STT
   - ⚠️ API key sẽ được MÃ HÓA khi import

📌 LƯU Ý QUAN TRỌNG:
   - Không xóa dòng tiêu đề
   - Giữ nguyên tên các cột
   - Lưu file dưới định dạng .xlsx
   - Upload lại file qua Telegram bot

📞 HỖ TRỢ:
   Liên hệ admin nếu cần giúp đỡ.
'''
    }
}


class ExcelConfigManager:
    """Quản lý export/import cấu hình qua Excel"""
    
    def __init__(self, user_manager=None, api_key_manager=None):
        """
        Args:
            user_manager: Instance của UserManager
            api_key_manager: Instance của APIKeyManager
        """
        self.user_manager = user_manager
        self.api_key_manager = api_key_manager
        
        # Template directory
        self.template_dir = Path("data/templates")
        self.template_dir.mkdir(parents=True, exist_ok=True)
        
        # Styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles"""
        if not EXCEL_AVAILABLE:
            return
            
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # ============================================================
    # GENERATE TEMPLATE
    # ============================================================
    def generate_blank_template(self) -> io.BytesIO:
        """
        Tạo template Excel trống cho user mới
        
        Returns:
            BytesIO buffer chứa file Excel
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas/openpyxl not installed")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        for sheet_name, sheet_config in TEMPLATE_SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)
            
            if sheet_name == 'Instructions':
                # Special handling for instructions
                self._add_instructions_sheet(ws, sheet_config['content'])
            else:
                self._add_data_sheet(ws, sheet_name, sheet_config)
        
        # Reorder sheets - Instructions first
        wb.move_sheet("Instructions", offset=-len(TEMPLATE_SHEETS)+1)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _add_instructions_sheet(self, ws, content: str):
        """Add instructions sheet"""
        ws.column_dimensions['A'].width = 80
        
        lines = content.strip().split('\n')
        for i, line in enumerate(lines, 1):
            cell = ws.cell(row=i, column=1, value=line)
            if line.startswith('📋') or line.startswith('📌'):
                cell.font = Font(bold=True, size=14, color="4F81BD")
            elif line.startswith(('1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣')):
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(size=11)
    
    def _add_data_sheet(self, ws, sheet_name: str, config: Dict):
        """Add data sheet with headers and sample data"""
        columns = config['columns']
        
        # Add description row
        ws.cell(row=1, column=1, value=f"📋 {config['description']}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.row_dimensions[1].height = 25
        
        # Add headers
        for col_idx, (col_id, col_name, sample, hint) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            
            # Set column width
            ws.column_dimensions[chr(64 + col_idx)].width = max(15, len(col_name) + 5)
        
        # Add hint row
        for col_idx, (col_id, col_name, sample, hint) in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=f"({hint})")
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = self.center_align
        
        # Add sample data row
        for col_idx, (col_id, col_name, sample, hint) in enumerate(columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=sample)
            cell.fill = self.alt_fill
            cell.alignment = self.left_align
            cell.border = self.thin_border
        
        # Add data validation for specific columns
        self._add_validations(ws, columns)
    
    def _add_validations(self, ws, columns):
        """Add data validation dropdowns"""
        col_map = {col[0]: idx for idx, col in enumerate(columns, 1)}
        
        # Speaking style validation
        if 'speaking_style' in col_map:
            dv = DataValidation(
                type="list",
                formula1='"friendly,professional,cute,playful,formal"',
                allow_blank=True
            )
            dv.add(ws.cell(row=4, column=col_map['speaking_style']))
            ws.add_data_validation(dv)
        
        # Language validation
        if 'primary_language' in col_map:
            dv = DataValidation(
                type="list",
                formula1='"vi,en,ja,zh,ko"',
                allow_blank=True
            )
            dv.add(ws.cell(row=4, column=col_map['primary_language']))
            ws.add_data_validation(dv)
        
        # Boolean validation
        for col_name in ['enabled', 'is_default']:
            if col_name in col_map:
                dv = DataValidation(
                    type="list",
                    formula1='"TRUE,FALSE"',
                    allow_blank=True
                )
                dv.add(ws.cell(row=4, column=col_map[col_name]))
                ws.add_data_validation(dv)
    
    # ============================================================
    # EXPORT USER CONFIG
    # ============================================================
    def export_user_config(self, user_id: int) -> io.BytesIO:
        """
        Export cấu hình của user ra Excel
        
        Args:
            user_id: Database user ID
            
        Returns:
            BytesIO buffer chứa file Excel
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas/openpyxl not installed")
        
        if not self.user_manager:
            raise ValueError("UserManager not provided")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Get user data
        user = self.user_manager.get_user_by_id(user_id)
        if not user:
            raise ValueError(f"User {user_id} not found")
        
        # Add sheets with user data
        self._export_personality_sheet(wb, user_id)
        self._export_knowledge_sheet(wb, user_id)
        self._export_contacts_sheet(wb, user_id)
        self._export_workflows_sheet(wb, user_id)
        self._export_api_keys_sheet(wb, user_id)
        self._add_instructions_sheet(wb.create_sheet("Instructions"), TEMPLATE_SHEETS['Instructions']['content'])
        
        # Reorder
        wb.move_sheet("Instructions", offset=-5)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _export_personality_sheet(self, wb, user_id: int):
        """Export personality config"""
        ws = wb.create_sheet("Personality")
        config = self.user_manager.get_personality_config(user_id) or {}
        
        columns = TEMPLATE_SHEETS['Personality']['columns']
        
        # Headers
        for col_idx, (col_id, col_name, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Data
        for col_idx, (col_id, _, default, _) in enumerate(columns, 1):
            value = config.get(col_id, default)
            ws.cell(row=2, column=col_idx, value=str(value) if value else '')
    
    def _export_knowledge_sheet(self, wb, user_id: int):
        """Export knowledge base"""
        ws = wb.create_sheet("Knowledge Base")
        
        # Get knowledge from user's collection
        knowledge_items = self.user_manager.get_user_knowledge(user_id) if hasattr(self.user_manager, 'get_user_knowledge') else []
        
        columns = TEMPLATE_SHEETS['Knowledge Base']['columns']
        
        # Headers
        for col_idx, (col_id, col_name, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 30
        
        # Data
        for row_idx, item in enumerate(knowledge_items, 2):
            for col_idx, (col_id, _, _, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(col_id, ''))
        
        # Add empty rows for new entries
        if len(knowledge_items) < 5:
            for row_idx in range(len(knowledge_items) + 2, 7):
                for col_idx in range(1, len(columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value='')
                    cell.border = self.thin_border
    
    def _export_contacts_sheet(self, wb, user_id: int):
        """Export contacts"""
        ws = wb.create_sheet("Contacts")
        
        # Get contacts from user's data
        contacts = self.user_manager.get_user_contacts(user_id) if hasattr(self.user_manager, 'get_user_contacts') else []
        
        columns = TEMPLATE_SHEETS['Contacts']['columns']
        
        # Headers
        for col_idx, (col_id, col_name, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 18
        
        # Data
        for row_idx, contact in enumerate(contacts, 2):
            for col_idx, (col_id, _, _, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=contact.get(col_id, ''))
    
    def _export_workflows_sheet(self, wb, user_id: int):
        """Export workflows"""
        ws = wb.create_sheet("Workflows")
        
        workflows = self.user_manager.get_user_workflows(user_id) if hasattr(self.user_manager, 'get_user_workflows') else []
        
        columns = TEMPLATE_SHEETS['Workflows']['columns']
        
        # Headers
        for col_idx, (col_id, col_name, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 18
        
        # Data
        for row_idx, wf in enumerate(workflows, 2):
            for col_idx, (col_id, _, _, _) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=wf.get(col_id, ''))
    
    def _export_api_keys_sheet(self, wb, user_id: int):
        """Export API keys (masked)"""
        ws = wb.create_sheet("API Keys")
        
        api_configs = self.user_manager.get_api_config(user_id) if self.user_manager else []
        
        columns = TEMPLATE_SHEETS['API Keys']['columns']
        
        # Headers
        for col_idx, (col_id, col_name, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            ws.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Warning row
        ws.cell(row=2, column=1, value="⚠️ API keys được che. Nhập key mới nếu muốn thay đổi.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        ws.cell(row=2, column=1).font = Font(italic=True, color="FF0000")
        
        # Data (masked)
        for row_idx, config in enumerate(api_configs, 3):
            for col_idx, (col_id, _, _, _) in enumerate(columns, 1):
                value = config.get(col_id, '')
                
                # Mask API key
                if col_id == 'api_key' and value:
                    if self.api_key_manager:
                        try:
                            _, decrypted = self.api_key_manager.decrypt_api_key(value)
                            value = self.api_key_manager.mask_api_key(decrypted) if decrypted else '***'
                        except:
                            value = '***encrypted***'
                    else:
                        value = '***'
                
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
    
    # ============================================================
    # IMPORT USER CONFIG
    # ============================================================
    def import_user_config(self, user_id: int, file_buffer: io.BytesIO) -> Dict[str, Any]:
        """
        Import cấu hình từ Excel file
        
        Args:
            user_id: Database user ID
            file_buffer: BytesIO buffer chứa file Excel
            
        Returns:
            Dict với kết quả import
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas/openpyxl not installed")
        
        if not self.user_manager:
            raise ValueError("UserManager not provided")
        
        results = {
            'success': True,
            'sheets_processed': [],
            'errors': [],
            'warnings': []
        }
        
        try:
            xl = pd.ExcelFile(file_buffer)
            
            for sheet_name in xl.sheet_names:
                if sheet_name == 'Instructions':
                    continue
                
                try:
                    df = pd.read_excel(xl, sheet_name=sheet_name, skiprows=2)  # Skip description and hint rows
                    
                    if sheet_name == 'Personality':
                        self._import_personality(user_id, df, results)
                    elif sheet_name == 'Knowledge Base':
                        self._import_knowledge(user_id, df, results)
                    elif sheet_name == 'Contacts':
                        self._import_contacts(user_id, df, results)
                    elif sheet_name == 'Workflows':
                        self._import_workflows(user_id, df, results)
                    elif sheet_name == 'API Keys':
                        self._import_api_keys(user_id, df, results)
                    
                    results['sheets_processed'].append(sheet_name)
                    
                except Exception as e:
                    results['errors'].append(f"{sheet_name}: {str(e)}")
                    logger.error(f"Error importing {sheet_name}: {e}")
        
        except Exception as e:
            results['success'] = False
            results['errors'].append(f"File error: {str(e)}")
            logger.error(f"Error reading Excel file: {e}")
        
        return results
    
    def _import_personality(self, user_id: int, df: pd.DataFrame, results: Dict):
        """Import personality config"""
        if df.empty:
            return
        
        row = df.iloc[0]
        config = {}
        
        col_mapping = {
            'Tên nhân vật': 'character_name',
            'Wake Word': 'wake_word',
            'Phong cách nói': 'speaking_style',
            'Ngôn ngữ': 'primary_language',
            'Temperature': 'temperature',
            'Độ dài trả lời': 'response_length',
        }
        
        for excel_col, db_col in col_mapping.items():
            if excel_col in row and pd.notna(row[excel_col]):
                config[db_col] = str(row[excel_col])
        
        if config:
            success = self.user_manager.update_personality_config(user_id, **config)
            if success:
                results['sheets_processed'].append('Personality')
            else:
                results['warnings'].append('Personality: Could not update')
    
    def _import_knowledge(self, user_id: int, df: pd.DataFrame, results: Dict):
        """Import knowledge base"""
        if df.empty:
            return
        
        # TODO: Implement knowledge import to ChromaDB
        count = len(df.dropna(how='all'))
        results['warnings'].append(f'Knowledge Base: {count} items found (import to ChromaDB not implemented)')
    
    def _import_contacts(self, user_id: int, df: pd.DataFrame, results: Dict):
        """Import contacts"""
        if df.empty:
            return
        
        # TODO: Implement contacts import
        count = len(df.dropna(how='all'))
        results['warnings'].append(f'Contacts: {count} items found (user contacts not implemented)')
    
    def _import_workflows(self, user_id: int, df: pd.DataFrame, results: Dict):
        """Import workflows"""
        if df.empty:
            return
        
        # TODO: Implement workflows import
        count = len(df.dropna(how='all'))
        results['warnings'].append(f'Workflows: {count} items found (user workflows not implemented)')
    
    def _import_api_keys(self, user_id: int, df: pd.DataFrame, results: Dict):
        """Import and encrypt API keys"""
        if df.empty:
            return
        
        if not self.api_key_manager:
            results['warnings'].append('API Keys: APIKeyManager not available')
            return
        
        col_mapping = {
            'Provider': 'provider',
            'Loại': 'type',
            'API Key': 'api_key',
            'API Base': 'api_base',
            'Model': 'model',
            'Mặc định': 'is_default',
        }
        
        imported = 0
        for _, row in df.iterrows():
            try:
                provider = row.get('Provider', '')
                api_key = row.get('API Key', '')
                
                if not provider or not api_key or api_key.startswith('***'):
                    continue
                
                # Encrypt API key
                encrypted = self.api_key_manager.encrypt_api_key(provider, api_key)
                
                # Save to database
                success = self.user_manager.save_api_config(
                    user_id=user_id,
                    provider_type=str(row.get('Loại', 'llm')),
                    provider_name=provider,
                    api_key=encrypted,
                    api_base=str(row.get('API Base', '')) if pd.notna(row.get('API Base')) else '',
                    model_name=str(row.get('Model', '')) if pd.notna(row.get('Model')) else '',
                    is_default=str(row.get('Mặc định', '')).upper() == 'TRUE'
                )
                
                if success:
                    imported += 1
                    
            except Exception as e:
                results['warnings'].append(f'API Keys row error: {e}')
        
        if imported > 0:
            results['sheets_processed'].append(f'API Keys ({imported} imported)')


# ============================================================
# FACTORY FUNCTION
# ============================================================
def get_excel_config_manager(user_manager=None, api_key_manager=None) -> ExcelConfigManager:
    """Factory function"""
    return ExcelConfigManager(user_manager, api_key_manager)


# ============================================================
# TEST
# ============================================================
if __name__ == '__main__':
    print("Testing Excel Config Manager...")
    
    manager = ExcelConfigManager()
    
    # Generate blank template
    buffer = manager.generate_blank_template()
    
    # Save to file for testing
    with open('data/templates/MeiLin_Config_Template.xlsx', 'wb') as f:
        f.write(buffer.read())
    
    print("✅ Template saved to data/templates/MeiLin_Config_Template.xlsx")
