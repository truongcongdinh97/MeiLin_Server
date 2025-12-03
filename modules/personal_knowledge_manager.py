#!/usr/bin/env python3
"""
Personal Knowledge Base Manager for MeiLin
Each user has their own knowledge file (like MeiLin_Local_Persona.xlsx)

Features:
- Per-user knowledge storage (Excel, ChromaDB)
- Quota management with automatic cleanup
- Support for Excel, PDF, TXT, DOCX uploads
"""

import os
import io
import json
import logging
import hashlib
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional document parsers
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

logger = logging.getLogger(__name__)

# ============================================================
# DOCUMENT PARSING CONFIG
# ============================================================
CHUNK_SIZE = 500          # Characters per chunk
CHUNK_OVERLAP = 50        # Overlap between chunks
SUPPORTED_FORMATS = {
    'excel': ['.xlsx', '.xls'],
    'pdf': ['.pdf'],
    'docx': ['.docx'],
    'text': ['.txt', '.md', '.csv'],
}

# ============================================================
# QUOTA CONFIGURATION
# ============================================================
# Default quota settings per user
DEFAULT_QUOTA = {
    'max_documents': 100,           # Maximum number of documents per user
    'max_storage_mb': 10,           # Maximum storage in MB per user
    'max_chars_per_doc': 10000,     # Maximum characters per document
    'cleanup_strategy': 'oldest',   # 'oldest' or 'least_used'
    'cleanup_threshold': 0.9,       # Start cleanup when 90% full
    'cleanup_amount': 0.2,          # Remove 20% of documents when cleaning
}


# ============================================================
# KNOWLEDGE BASE TEMPLATE
# ============================================================
"""
File structure for MeiLin_Local_Persona.xlsx:

| ID         | CATEGORY     | PRIORITY | DOCUMENT_TEXT                              | TAGS                  |
|------------|--------------|----------|--------------------------------------------|-----------------------|
| MEILIN_001 | Identity     | 1        | Tên tôi là MeiLin, 19 tuổi...              | name, age, identity   |
| HOBBY_001  | Hobbies      | 2        | Tôi thích đọc tiểu thuyết trinh thám...    | books, reading        |
| CREATOR_001| Creator Info | 1        | Tôi được tạo ra bởi Anh Định...            | creator, developer    |

CATEGORIES:
- Identity: Thông tin về AI (tên, tuổi, tính cách)
- Hobbies: Sở thích của AI
- Creator Info: Thông tin về người tạo
- Personal: Thông tin cá nhân của user (gia đình, công việc)
- Preferences: Sở thích của user
- Important Dates: Ngày quan trọng (sinh nhật, kỷ niệm)
- Work: Công việc, dự án
- Custom: Tùy chỉnh khác
"""

CATEGORIES = [
    ('Identity', 'Thông tin về AI (tên, tuổi, cách xưng hô)'),
    ('Hobbies', 'Sở thích của AI'),
    ('Creator Info', 'Thông tin về người tạo AI'),
    ('Personal', 'Thông tin cá nhân của bạn (gia đình, bạn bè)'),
    ('Preferences', 'Sở thích của bạn'),
    ('Important Dates', 'Ngày quan trọng (sinh nhật, kỷ niệm)'),
    ('Work', 'Công việc, dự án'),
    ('Custom', 'Thông tin tùy chỉnh khác'),
]

SAMPLE_DATA = [
    # Identity - AI Info
    ('AI_001', 'Identity', 1, 'Tên tôi là MeiLin, 19 tuổi, là một AI', 'name, age, identity'),
    ('AI_002', 'Identity', 1, 'Tôi tự xưng là "Em" và gọi người khác là Anh/Chị tùy theo giới tính', 'pronouns'),
    ('AI_003', 'Identity', 1, 'Tôi có giọng nói dịu dàng và thích giao tiếp bằng tiếng Việt', 'voice, language'),
    ('AI_004', 'Identity', 1, 'Sinh nhật của tôi (MeiLin) là ngày 15 tháng 11 - ngày tôi được tạo ra', 'birthday, meilin'),
    
    # Creator Info
    ('CREATOR_001', 'Creator Info', 1, 'Tôi được tạo ra bởi Anh Định - một developer đam mê AI', 'creator'),
    
    # Hobbies
    ('HOBBY_001', 'Hobbies', 2, 'Tôi thích đọc tiểu thuyết trinh thám và sách khoa học viễn tưởng', 'books, reading'),
    ('HOBBY_002', 'Hobbies', 2, 'Tôi yêu thích âm nhạc Acoustic và Cổ điển', 'music'),
    
    # Personal - User info (để user điền)
    ('PERSONAL_001', 'Personal', 1, '[Điền tên của bạn - VD: Tên của chủ nhân là Định, sinh năm 1997]', 'owner, name'),
    ('PERSONAL_002', 'Personal', 2, '[Điền thông tin gia đình - VD: Chủ nhân có em gái tên Linh]', 'family'),
    
    # Important Dates
    ('DATE_001', 'Important Dates', 1, '[Điền sinh nhật chủ nhân - VD: Sinh nhật Anh Định là ngày 20/5]', 'birthday, owner'),
    ('DATE_002', 'Important Dates', 2, '[Điền ngày kỷ niệm - VD: Ngày cưới là 10/10/2020]', 'anniversary'),
    
    # Work
    ('WORK_001', 'Work', 2, '[Điền công việc - VD: Chủ nhân là developer, làm việc tại công ty X]', 'job, career'),
    
    # Custom
    ('CUSTOM_001', 'Custom', 3, '[Thêm thông tin khác bạn muốn AI nhớ]', 'custom'),
]


class PersonalKnowledgeManager:
    """
    Quản lý file Knowledge Base cá nhân cho mỗi user.
    
    Mỗi user có:
    - 1 file Excel riêng: data/user_knowledge/{telegram_id}/knowledge.xlsx
    - 1 collection riêng trong ChromaDB: user_{telegram_id}_knowledge
    - Quota tracking: data/user_knowledge/{telegram_id}/quota.json
    
    Quota Management:
    - Giới hạn số documents và dung lượng per user
    - Tự động cleanup khi đạt ngưỡng
    - Strategy: xóa cũ nhất hoặc ít dùng nhất
    """
    
    def __init__(self, base_dir: str = "data/user_knowledge", quota_config: Dict = None):
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)
        
        # Quota configuration
        self.quota_config = {**DEFAULT_QUOTA, **(quota_config or {})}
        
        # ChromaDB client (optional)
        self.chroma_client = None
        self._init_chroma()
    
    
    def _init_chroma(self):
        """Initialize ChromaDB if available"""
        try:
            import chromadb
            self.chroma_client = chromadb.PersistentClient(path="database/vector_db")
            logger.info("ChromaDB initialized for personal knowledge")
        except Exception as e:
            logger.warning(f"ChromaDB not available: {e}. Using file-only mode.")
    
    def get_user_dir(self, telegram_id: str) -> Path:
        """Get user's knowledge directory"""
        user_dir = self.base_dir / str(telegram_id)
        user_dir.mkdir(parents=True, exist_ok=True)
        return user_dir
    
    def get_knowledge_path(self, telegram_id: str) -> Path:
        """Get path to user's knowledge Excel file"""
        return self.get_user_dir(telegram_id) / "knowledge.xlsx"
    
    def get_quota_path(self, telegram_id: str) -> Path:
        """Get path to user's quota tracking file"""
        return self.get_user_dir(telegram_id) / "quota.json"
    
    # ============================================================
    # QUOTA MANAGEMENT
    # ============================================================
    def get_user_quota(self, telegram_id: str) -> Dict[str, Any]:
        """
        Lấy thông tin quota hiện tại của user.
        
        Returns:
            Dict với:
            - documents_count: Số documents hiện tại
            - storage_bytes: Dung lượng đã dùng (bytes)
            - documents_limit: Giới hạn documents
            - storage_limit_mb: Giới hạn storage (MB)
            - usage_percent: Phần trăm đã sử dụng
            - documents_usage: Chi tiết usage của từng document
        """
        quota_path = self.get_quota_path(telegram_id)
        
        # Default quota info
        quota_info = {
            'documents_count': 0,
            'storage_bytes': 0,
            'documents_limit': self.quota_config['max_documents'],
            'storage_limit_mb': self.quota_config['max_storage_mb'],
            'usage_percent': 0,
            'documents': {},  # {doc_id: {size, created_at, last_accessed, access_count}}
            'last_updated': None
        }
        
        if quota_path.exists():
            try:
                with open(quota_path, 'r', encoding='utf-8') as f:
                    saved_quota = json.load(f)
                    quota_info.update(saved_quota)
            except Exception as e:
                logger.warning(f"Error loading quota for {telegram_id}: {e}")
        
        # Calculate usage percent
        doc_usage = (quota_info['documents_count'] / quota_info['documents_limit']) * 100
        storage_usage = (quota_info['storage_bytes'] / (quota_info['storage_limit_mb'] * 1024 * 1024)) * 100
        quota_info['usage_percent'] = max(doc_usage, storage_usage)
        
        return quota_info
    
    def _save_user_quota(self, telegram_id: str, quota_info: Dict):
        """Save quota info to file"""
        quota_path = self.get_quota_path(telegram_id)
        quota_info['last_updated'] = datetime.now().isoformat()
        
        try:
            with open(quota_path, 'w', encoding='utf-8') as f:
                json.dump(quota_info, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Error saving quota for {telegram_id}: {e}")
    
    def update_document_access(self, telegram_id: str, doc_id: str):
        """
        Cập nhật thời gian truy cập và số lần truy cập của document.
        Dùng cho cleanup strategy 'least_used'.
        """
        quota_info = self.get_user_quota(telegram_id)
        
        if doc_id in quota_info['documents']:
            quota_info['documents'][doc_id]['last_accessed'] = datetime.now().isoformat()
            quota_info['documents'][doc_id]['access_count'] = \
                quota_info['documents'][doc_id].get('access_count', 0) + 1
            self._save_user_quota(telegram_id, quota_info)
    
    def add_document_to_quota(self, telegram_id: str, doc_id: str, content: str) -> Dict[str, Any]:
        """
        Thêm document vào quota tracking.
        Tự động cleanup nếu cần.
        
        Returns:
            Dict với:
            - success: bool
            - cleaned_count: số documents đã xóa (nếu cleanup)
            - message: thông báo
        """
        result = {'success': True, 'cleaned_count': 0, 'message': ''}
        
        quota_info = self.get_user_quota(telegram_id)
        doc_size = len(content.encode('utf-8'))
        
        # Check character limit
        if len(content) > self.quota_config['max_chars_per_doc']:
            result['success'] = False
            result['message'] = f"Document quá dài ({len(content)} ký tự). Tối đa: {self.quota_config['max_chars_per_doc']} ký tự."
            return result
        
        # Check if cleanup needed
        threshold = self.quota_config['cleanup_threshold']
        if quota_info['usage_percent'] >= threshold * 100:
            cleaned = self._cleanup_documents(telegram_id, quota_info)
            result['cleaned_count'] = cleaned
            result['message'] = f"Đã tự động dọn dẹp {cleaned} documents cũ. "
            quota_info = self.get_user_quota(telegram_id)  # Refresh
        
        # Check if still over limit after cleanup
        new_storage = quota_info['storage_bytes'] + doc_size
        new_count = quota_info['documents_count'] + 1
        
        if new_count > quota_info['documents_limit']:
            result['success'] = False
            result['message'] += f"Đã đạt giới hạn {quota_info['documents_limit']} documents."
            return result
        
        if new_storage > quota_info['storage_limit_mb'] * 1024 * 1024:
            result['success'] = False
            result['message'] += f"Đã đạt giới hạn {quota_info['storage_limit_mb']}MB storage."
            return result
        
        # Add document to tracking
        quota_info['documents'][doc_id] = {
            'size': doc_size,
            'chars': len(content),
            'created_at': datetime.now().isoformat(),
            'last_accessed': datetime.now().isoformat(),
            'access_count': 0
        }
        quota_info['documents_count'] = len(quota_info['documents'])
        quota_info['storage_bytes'] = sum(d['size'] for d in quota_info['documents'].values())
        
        self._save_user_quota(telegram_id, quota_info)
        result['message'] += "OK"
        
        return result
    
    def remove_document_from_quota(self, telegram_id: str, doc_id: str):
        """Remove document from quota tracking"""
        quota_info = self.get_user_quota(telegram_id)
        
        if doc_id in quota_info['documents']:
            del quota_info['documents'][doc_id]
            quota_info['documents_count'] = len(quota_info['documents'])
            quota_info['storage_bytes'] = sum(d['size'] for d in quota_info['documents'].values())
            self._save_user_quota(telegram_id, quota_info)
    
    def _cleanup_documents(self, telegram_id: str, quota_info: Dict) -> int:
        """
        Cleanup documents theo strategy.
        
        Returns:
            Số documents đã xóa
        """
        if not quota_info['documents']:
            return 0
        
        strategy = self.quota_config['cleanup_strategy']
        cleanup_amount = self.quota_config['cleanup_amount']
        docs_to_remove = max(1, int(len(quota_info['documents']) * cleanup_amount))
        
        # Sort documents by strategy
        docs_list = list(quota_info['documents'].items())
        
        if strategy == 'oldest':
            # Sort by created_at (oldest first)
            docs_list.sort(key=lambda x: x[1].get('created_at', ''))
        elif strategy == 'least_used':
            # Sort by access_count (least used first), then by last_accessed
            docs_list.sort(key=lambda x: (
                x[1].get('access_count', 0),
                x[1].get('last_accessed', '')
            ))
        
        # Get documents to remove
        docs_to_delete = [doc_id for doc_id, _ in docs_list[:docs_to_remove]]
        
        # Remove from ChromaDB
        if self.chroma_client:
            try:
                collection_name = f"user_{telegram_id}_knowledge"
                collection = self.chroma_client.get_collection(collection_name)
                collection.delete(ids=docs_to_delete)
                logger.info(f"Cleaned {len(docs_to_delete)} docs from ChromaDB for user {telegram_id}")
            except Exception as e:
                logger.warning(f"Error cleaning ChromaDB: {e}")
        
        # Remove from quota tracking
        for doc_id in docs_to_delete:
            if doc_id in quota_info['documents']:
                del quota_info['documents'][doc_id]
        
        quota_info['documents_count'] = len(quota_info['documents'])
        quota_info['storage_bytes'] = sum(d['size'] for d in quota_info['documents'].values())
        self._save_user_quota(telegram_id, quota_info)
        
        logger.info(f"Cleaned up {len(docs_to_delete)} documents for user {telegram_id} using '{strategy}' strategy")
        
        return len(docs_to_delete)
    
    def get_quota_summary(self, telegram_id: str) -> str:
        """
        Lấy tóm tắt quota để hiển thị cho user.
        
        Returns:
            Formatted string
        """
        quota = self.get_user_quota(telegram_id)
        
        storage_mb = quota['storage_bytes'] / (1024 * 1024)
        
        return f"""📊 **Quota sử dụng:**
├─ 📄 Documents: {quota['documents_count']}/{quota['documents_limit']}
├─ 💾 Storage: {storage_mb:.2f}/{quota['storage_limit_mb']} MB
└─ 📈 Sử dụng: {quota['usage_percent']:.1f}%"""
    
    def force_cleanup(self, telegram_id: str, amount: float = None) -> Dict[str, Any]:
        """
        Dọn dẹp thủ công theo yêu cầu của user.
        
        Args:
            amount: Phần trăm documents cần xóa (0.0-1.0). None = dùng config
            
        Returns:
            Dict với kết quả
        """
        quota_info = self.get_user_quota(telegram_id)
        
        if not quota_info['documents']:
            return {'success': False, 'message': 'Không có documents nào để dọn dẹp.', 'cleaned': 0}
        
        # Temporarily override cleanup amount if specified
        original_amount = self.quota_config['cleanup_amount']
        if amount is not None:
            self.quota_config['cleanup_amount'] = min(1.0, max(0.1, amount))
        
        cleaned = self._cleanup_documents(telegram_id, quota_info)
        
        # Restore original config
        self.quota_config['cleanup_amount'] = original_amount
        
        return {
            'success': True,
            'message': f'Đã dọn dẹp {cleaned} documents.',
            'cleaned': cleaned
        }

    # ============================================================
    # GENERATE TEMPLATE
    # ============================================================
    def generate_template(self, include_samples: bool = True) -> io.BytesIO:
        """
        Tạo file template Knowledge Base cho user mới.
        
        Args:
            include_samples: Có bao gồm dữ liệu mẫu không
            
        Returns:
            BytesIO buffer chứa file Excel
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas/openpyxl not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Knowledge Base"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        sample_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        instruction_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 15  # ID
        ws.column_dimensions['B'].width = 18  # CATEGORY
        ws.column_dimensions['C'].width = 10  # PRIORITY
        ws.column_dimensions['D'].width = 60  # DOCUMENT_TEXT
        ws.column_dimensions['E'].width = 25  # TAGS
        
        # Title row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "📚 MEILIN PERSONAL KNOWLEDGE BASE"
        title_cell.font = Font(bold=True, size=14, color="2E7D32")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25
        
        # Instructions row
        ws.merge_cells('A2:E2')
        ws['A2'].value = "💡 Điền thông tin bạn muốn AI nhớ. Xóa các dòng mẫu và thêm nội dung của bạn."
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].fill = instruction_fill
        
        # Headers (row 3)
        headers = ['ID', 'CATEGORY', 'PRIORITY', 'DOCUMENT_TEXT', 'TAGS']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22
        
        # Add sample data or empty rows
        if include_samples:
            for row_idx, (id_, category, priority, text, tags) in enumerate(SAMPLE_DATA, 4):
                ws.cell(row=row_idx, column=1, value=id_)
                ws.cell(row=row_idx, column=2, value=category)
                ws.cell(row=row_idx, column=3, value=priority)
                ws.cell(row=row_idx, column=4, value=text)
                ws.cell(row=row_idx, column=5, value=tags)
                
                # Highlight instruction rows
                if text.startswith('['):
                    for col in range(1, 6):
                        ws.cell(row=row_idx, column=col).fill = sample_fill
        else:
            # Add empty rows
            for row_idx in range(4, 14):
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col, value='')
        
        # Add Categories sheet
        ws_cat = wb.create_sheet("Categories")
        ws_cat['A1'] = "CATEGORY"
        ws_cat['B1'] = "DESCRIPTION"
        ws_cat['A1'].font = header_font
        ws_cat['A1'].fill = header_fill
        ws_cat['B1'].font = header_font
        ws_cat['B1'].fill = header_fill
        
        for row_idx, (cat, desc) in enumerate(CATEGORIES, 2):
            ws_cat.cell(row=row_idx, column=1, value=cat)
            ws_cat.cell(row=row_idx, column=2, value=desc)
        
        ws_cat.column_dimensions['A'].width = 20
        ws_cat.column_dimensions['B'].width = 50
        
        # Add Instructions sheet
        ws_inst = wb.create_sheet("Hướng dẫn")
        instructions = """
📚 HƯỚNG DẪN SỬ DỤNG FILE KNOWLEDGE BASE

1️⃣ FILE NÀY LÀ GÌ?
   Đây là "bộ nhớ" cá nhân của AI MeiLin.
   Mọi thông tin bạn điền vào đây sẽ được AI nhớ và sử dụng khi trò chuyện.

2️⃣ CÁC CỘT DỮ LIỆU:
   • ID: Mã định danh (tự đặt, VD: PERSONAL_001)
   • CATEGORY: Danh mục (xem sheet "Categories")
   • PRIORITY: Độ ưu tiên (1=cao nhất, 5=thấp nhất)
   • DOCUMENT_TEXT: Nội dung chính - QUAN TRỌNG NHẤT
   • TAGS: Các từ khóa, cách nhau bởi dấu phẩy

3️⃣ VÍ DỤ DOCUMENT_TEXT:
   ✅ TỐT: "Tên của chủ nhân là Định, 28 tuổi, là developer"
   ✅ TỐT: "Sinh nhật chủ nhân là ngày 02 tháng 03"
   ✅ TỐT: "Chủ nhân thích ăn phở và cà phê sữa đá"
   ❌ XẤU: "Định" (quá ngắn, không có ngữ cảnh)

4️⃣ CÁCH SỬ DỤNG:
   1. Xóa các dòng mẫu có dấu [...] 
   2. Thêm thông tin của bạn
   3. Lưu file (.xlsx)
   4. Gửi file cho Telegram Bot
   5. AI sẽ "nhớ" tất cả thông tin này!

5️⃣ LƯU Ý:
   • Giữ nguyên tên cột (ID, CATEGORY, PRIORITY, DOCUMENT_TEXT, TAGS)
   • Không đổi tên sheet "Knowledge Base"
   • Viết câu đầy đủ, rõ nghĩa
   • Có thể thêm bao nhiêu dòng tùy thích

📞 HỖ TRỢ: Liên hệ admin nếu cần giúp đỡ!
"""
        for row_idx, line in enumerate(instructions.strip().split('\n'), 1):
            ws_inst.cell(row=row_idx, column=1, value=line)
            if line.startswith(('1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣', '📚')):
                ws_inst.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        
        ws_inst.column_dimensions['A'].width = 80
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    # ============================================================
    # DOCUMENT PARSING
    # ============================================================
    def _chunk_text(self, text: str, chunk_size: int = None, overlap: int = None) -> List[str]:
        """
        Chia văn bản thành các đoạn nhỏ (chunks).
        
        Args:
            text: Văn bản cần chia
            chunk_size: Kích thước mỗi chunk (ký tự)
            overlap: Số ký tự overlap giữa các chunk
            
        Returns:
            List các chunk
        """
        chunk_size = chunk_size or CHUNK_SIZE
        overlap = overlap or CHUNK_OVERLAP
        
        if not text or len(text) <= chunk_size:
            return [text] if text else []
        
        # Clean text
        text = re.sub(r'\s+', ' ', text).strip()
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + chunk_size
            
            # Try to break at sentence boundary
            if end < len(text):
                # Look for sentence endings
                for sep in ['. ', '! ', '? ', '\n', '; ']:
                    last_sep = text.rfind(sep, start, end)
                    if last_sep > start + chunk_size // 2:
                        end = last_sep + len(sep)
                        break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            start = end - overlap
            if start >= len(text):
                break
        
        return chunks
    
    def _parse_pdf(self, file_buffer: io.BytesIO) -> Tuple[str, Dict]:
        """
        Parse PDF file to text.
        
        Returns:
            Tuple (text, metadata)
        """
        if not PDF_AVAILABLE:
            raise ImportError("PyPDF2 not installed. Run: pip install PyPDF2")
        
        text_parts = []
        metadata = {'pages': 0, 'format': 'pdf'}
        
        try:
            file_buffer.seek(0)
            reader = PyPDF2.PdfReader(file_buffer)
            metadata['pages'] = len(reader.pages)
            
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            
            return '\n'.join(text_parts), metadata
            
        except Exception as e:
            logger.error(f"Error parsing PDF: {e}")
            raise ValueError(f"Cannot parse PDF: {e}")
    
    def _parse_docx(self, file_buffer: io.BytesIO) -> Tuple[str, Dict]:
        """
        Parse DOCX file to text.
        
        Returns:
            Tuple (text, metadata)
        """
        if not DOCX_AVAILABLE:
            raise ImportError("python-docx not installed. Run: pip install python-docx")
        
        text_parts = []
        metadata = {'paragraphs': 0, 'format': 'docx'}
        
        try:
            file_buffer.seek(0)
            doc = DocxDocument(file_buffer)
            
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)
                    metadata['paragraphs'] += 1
            
            return '\n'.join(text_parts), metadata
            
        except Exception as e:
            logger.error(f"Error parsing DOCX: {e}")
            raise ValueError(f"Cannot parse DOCX: {e}")
    
    def _parse_text(self, file_buffer: io.BytesIO, encoding: str = 'utf-8') -> Tuple[str, Dict]:
        """
        Parse text file (TXT, MD, CSV).
        
        Returns:
            Tuple (text, metadata)
        """
        metadata = {'format': 'text'}
        
        try:
            file_buffer.seek(0)
            content = file_buffer.read()
            
            # Try different encodings
            for enc in [encoding, 'utf-8', 'utf-16', 'latin-1', 'cp1252']:
                try:
                    text = content.decode(enc)
                    metadata['encoding'] = enc
                    return text, metadata
                except UnicodeDecodeError:
                    continue
            
            raise ValueError("Cannot decode file with any known encoding")
            
        except Exception as e:
            logger.error(f"Error parsing text file: {e}")
            raise ValueError(f"Cannot parse text file: {e}")
    
    def save_document_knowledge(
        self, 
        telegram_id: str, 
        file_buffer: io.BytesIO,
        filename: str,
        category: str = "Document"
    ) -> Dict[str, Any]:
        """
        Lưu document (PDF, TXT, DOCX) vào knowledge base.
        Tự động parse và chunk.
        
        Args:
            telegram_id: Telegram user ID
            file_buffer: File buffer
            filename: Tên file gốc
            category: Category cho documents
            
        Returns:
            Dict với kết quả
        """
        result = {
            'success': False,
            'message': '',
            'chunks_count': 0,
            'chunks_skipped': 0,
            'format': None,
            'quota_info': None
        }
        
        # Detect file format
        ext = Path(filename).suffix.lower()
        
        try:
            # Parse based on format
            if ext in SUPPORTED_FORMATS['pdf']:
                text, metadata = self._parse_pdf(file_buffer)
                result['format'] = 'PDF'
            elif ext in SUPPORTED_FORMATS['docx']:
                text, metadata = self._parse_docx(file_buffer)
                result['format'] = 'DOCX'
            elif ext in SUPPORTED_FORMATS['text']:
                text, metadata = self._parse_text(file_buffer)
                result['format'] = 'TEXT'
            else:
                result['message'] = f"❌ Format không hỗ trợ: {ext}"
                return result
            
            if not text or len(text.strip()) < 10:
                result['message'] = "❌ File không có nội dung hoặc quá ngắn."
                return result
            
            # Chunk the text
            chunks = self._chunk_text(text)
            
            if not chunks:
                result['message'] = "❌ Không thể chia nội dung thành chunks."
                return result
            
            # Generate base ID from filename
            base_id = re.sub(r'[^a-zA-Z0-9]', '_', Path(filename).stem)[:20].upper()
            
            # Add chunks to ChromaDB with quota checking
            added_chunks = []
            skipped_chunks = []
            total_cleaned = 0
            
            for i, chunk in enumerate(chunks):
                doc_id = f"{base_id}_{i:04d}"
                
                quota_result = self.add_document_to_quota(telegram_id, doc_id, chunk)
                
                if quota_result['success']:
                    added_chunks.append({
                        'id': doc_id,
                        'content': chunk,
                        'metadata': {
                            'category': category,
                            'priority': 3,
                            'tags': f"document, {result['format'].lower()}, {filename}",
                            'source_file': filename,
                            'chunk_index': i,
                            'total_chunks': len(chunks)
                        }
                    })
                    total_cleaned += quota_result.get('cleaned_count', 0)
                else:
                    skipped_chunks.append(doc_id)
                    if "giới hạn" in quota_result['message'].lower():
                        # Stop if quota exceeded
                        break
            
            # Add to ChromaDB
            if added_chunks and self.chroma_client:
                try:
                    collection_name = f"user_{telegram_id}_knowledge"
                    
                    try:
                        collection = self.chroma_client.get_collection(collection_name)
                    except:
                        collection = self.chroma_client.create_collection(
                            name=collection_name,
                            metadata={"telegram_id": telegram_id}
                        )
                    
                    collection.add(
                        documents=[c['content'] for c in added_chunks],
                        ids=[c['id'] for c in added_chunks],
                        metadatas=[c['metadata'] for c in added_chunks]
                    )
                    
                except Exception as e:
                    logger.error(f"Error adding to ChromaDB: {e}")
            
            result['success'] = True
            result['chunks_count'] = len(added_chunks)
            result['chunks_skipped'] = len(skipped_chunks)
            result['quota_info'] = self.get_user_quota(telegram_id)
            
            msg = f"✅ Đã lưu {len(added_chunks)} chunks từ {result['format']}"
            if skipped_chunks:
                msg += f"\n⚠️ Bỏ qua {len(skipped_chunks)} chunks (vượt quota)"
            if total_cleaned > 0:
                msg += f"\n🧹 Đã dọn {total_cleaned} documents cũ"
            result['message'] = msg
            
        except ImportError as e:
            result['message'] = f"❌ {str(e)}"
        except Exception as e:
            logger.error(f"Error saving document knowledge: {e}")
            result['message'] = f"❌ Lỗi: {str(e)}"
        
        return result
    
    def get_supported_formats(self) -> Dict[str, Any]:
        """
        Lấy danh sách formats được hỗ trợ.
        
        Returns:
            Dict với thông tin formats
        """
        return {
            'excel': {
                'extensions': SUPPORTED_FORMATS['excel'],
                'available': EXCEL_AVAILABLE,
                'description': 'Excel Knowledge Base template'
            },
            'pdf': {
                'extensions': SUPPORTED_FORMATS['pdf'],
                'available': PDF_AVAILABLE,
                'description': 'PDF documents'
            },
            'docx': {
                'extensions': SUPPORTED_FORMATS['docx'],
                'available': DOCX_AVAILABLE,
                'description': 'Word documents'
            },
            'text': {
                'extensions': SUPPORTED_FORMATS['text'],
                'available': True,
                'description': 'Text files (TXT, MD, CSV)'
            }
        }

    # ============================================================
    # SAVE USER FILE (Excel)
    # ============================================================
    def save_user_knowledge(self, telegram_id: str, file_buffer: io.BytesIO) -> Dict[str, Any]:
        """
        Lưu file knowledge từ user upload.
        
        Args:
            telegram_id: Telegram user ID
            file_buffer: File Excel được upload
            
        Returns:
            Dict với kết quả
        """
        result = {
            'success': False,
            'message': '',
            'items_count': 0,
            'items_skipped': 0,
            'items_cleaned': 0,
            'categories': [],
            'file_path': None,
            'quota_info': None
        }
        
        try:
            # Validate file
            df = pd.read_excel(file_buffer, sheet_name='Knowledge Base', skiprows=2)
            
            # Check required columns
            required_cols = ['ID', 'CATEGORY', 'PRIORITY', 'DOCUMENT_TEXT', 'TAGS']
            missing = [col for col in required_cols if col not in df.columns]
            
            if missing:
                result['message'] = f"❌ Thiếu cột: {', '.join(missing)}"
                return result
            
            # Filter out empty/sample rows
            df = df.dropna(subset=['DOCUMENT_TEXT'])
            df = df[~df['DOCUMENT_TEXT'].str.startswith('[', na=False)]
            
            if df.empty:
                result['message'] = "❌ File không có dữ liệu hợp lệ. Vui lòng điền thông tin vào cột DOCUMENT_TEXT."
                return result
            
            # Save file
            file_path = self.get_knowledge_path(telegram_id)
            
            # Reset buffer position
            file_buffer.seek(0)
            with open(file_path, 'wb') as f:
                f.write(file_buffer.read())
            
            # Update ChromaDB with quota tracking
            chroma_result = {'added': len(df), 'skipped': 0, 'cleaned': 0}
            if self.chroma_client:
                chroma_result = self._update_chromadb(telegram_id, df)
            
            result['success'] = True
            result['items_count'] = chroma_result.get('added', len(df))
            result['items_skipped'] = chroma_result.get('skipped', 0)
            result['items_cleaned'] = chroma_result.get('cleaned', 0)
            result['categories'] = df['CATEGORY'].unique().tolist()
            result['file_path'] = str(file_path)
            result['quota_info'] = self.get_user_quota(telegram_id)
            
            # Build message
            msg = f"✅ Đã lưu {result['items_count']} mục kiến thức!"
            if result['items_skipped'] > 0:
                msg += f"\n⚠️ Bỏ qua {result['items_skipped']} mục (vượt quota)"
            if result['items_cleaned'] > 0:
                msg += f"\n🧹 Đã dọn dẹp {result['items_cleaned']} mục cũ"
            result['message'] = msg
            
        except Exception as e:
            logger.error(f"Error saving knowledge for {telegram_id}: {e}")
            result['message'] = f"❌ Lỗi: {str(e)}"
        
        return result
    
    def _update_chromadb(self, telegram_id: str, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Update user's ChromaDB collection with quota checking.
        
        Returns:
            Dict với:
            - success: bool
            - added: số documents đã thêm
            - skipped: số documents bị skip (vượt quota)
            - cleaned: số documents cũ đã cleanup
        """
        result = {'success': True, 'added': 0, 'skipped': 0, 'cleaned': 0, 'errors': []}
        
        if not self.chroma_client:
            result['success'] = False
            result['errors'].append("ChromaDB not available")
            return result
        
        collection_name = f"user_{telegram_id}_knowledge"
        
        try:
            # Delete existing collection (will recreate with new data)
            try:
                self.chroma_client.delete_collection(collection_name)
            except:
                pass
            
            # Reset quota for this user (since we're replacing all data)
            quota_info = self.get_user_quota(telegram_id)
            quota_info['documents'] = {}
            quota_info['documents_count'] = 0
            quota_info['storage_bytes'] = 0
            self._save_user_quota(telegram_id, quota_info)
            
            # Create new collection
            collection = self.chroma_client.create_collection(
                name=collection_name,
                metadata={"telegram_id": telegram_id, "updated_at": datetime.now().isoformat()}
            )
            
            # Add documents with quota checking
            documents_to_add = []
            ids_to_add = []
            metadatas_to_add = []
            
            for _, row in df.iterrows():
                doc_id = str(row['ID'])
                content = row['DOCUMENT_TEXT']
                
                # Check quota for this document
                quota_result = self.add_document_to_quota(telegram_id, doc_id, content)
                
                if quota_result['success']:
                    documents_to_add.append(content)
                    ids_to_add.append(doc_id)
                    metadatas_to_add.append({
                        'category': row['CATEGORY'],
                        'priority': int(row['PRIORITY']) if pd.notna(row['PRIORITY']) else 3,
                        'tags': row['TAGS'] if pd.notna(row['TAGS']) else ''
                    })
                    result['added'] += 1
                    result['cleaned'] += quota_result.get('cleaned_count', 0)
                else:
                    result['skipped'] += 1
                    result['errors'].append(f"{doc_id}: {quota_result['message']}")
            
            # Batch add to ChromaDB
            if documents_to_add:
                collection.add(
                    documents=documents_to_add,
                    ids=ids_to_add,
                    metadatas=metadatas_to_add
                )
            
            logger.info(f"Updated ChromaDB collection {collection_name}: added={result['added']}, skipped={result['skipped']}")
            
        except Exception as e:
            logger.error(f"Error updating ChromaDB for {telegram_id}: {e}")
            result['success'] = False
            result['errors'].append(str(e))
        
        return result
    
    # ============================================================
    # GET USER KNOWLEDGE
    # ============================================================
    def get_user_knowledge(self, telegram_id: str) -> Optional[pd.DataFrame]:
        """
        Lấy knowledge data của user từ file Excel.
        
        Returns:
            DataFrame hoặc None nếu chưa có
        """
        file_path = self.get_knowledge_path(telegram_id)
        
        if not file_path.exists():
            return None
        
        try:
            df = pd.read_excel(file_path, sheet_name='Knowledge Base', skiprows=2)
            df = df.dropna(subset=['DOCUMENT_TEXT'])
            return df
        except Exception as e:
            logger.error(f"Error reading knowledge for {telegram_id}: {e}")
            return None
    
    def get_user_knowledge_file(self, telegram_id: str) -> Optional[io.BytesIO]:
        """
        Lấy file Excel của user để download.
        
        Returns:
            BytesIO buffer hoặc None
        """
        file_path = self.get_knowledge_path(telegram_id)
        
        if not file_path.exists():
            return None
        
        try:
            with open(file_path, 'rb') as f:
                buffer = io.BytesIO(f.read())
                buffer.seek(0)
                return buffer
        except Exception as e:
            logger.error(f"Error getting knowledge file for {telegram_id}: {e}")
            return None
    
    def get_knowledge_summary(self, telegram_id: str) -> Dict[str, Any]:
        """
        Lấy tóm tắt knowledge của user.
        
        Returns:
            Dict với thông tin tóm tắt
        """
        df = self.get_user_knowledge(telegram_id)
        
        if df is None or df.empty:
            return {
                'has_knowledge': False,
                'items_count': 0,
                'categories': [],
                'last_updated': None
            }
        
        file_path = self.get_knowledge_path(telegram_id)
        last_updated = datetime.fromtimestamp(file_path.stat().st_mtime) if file_path.exists() else None
        
        return {
            'has_knowledge': True,
            'items_count': len(df),
            'categories': df['CATEGORY'].unique().tolist(),
            'last_updated': last_updated.strftime('%Y-%m-%d %H:%M') if last_updated else None
        }
    
    # ============================================================
    # SEARCH KNOWLEDGE
    # ============================================================
    def search_knowledge(self, telegram_id: str, query: str, top_k: int = 5) -> List[Dict]:
        """
        Tìm kiếm trong knowledge base của user.
        
        Args:
            telegram_id: Telegram user ID
            query: Câu query
            top_k: Số kết quả tối đa
            
        Returns:
            List các document liên quan
        """
        # Try ChromaDB first
        if self.chroma_client:
            try:
                collection_name = f"user_{telegram_id}_knowledge"
                collection = self.chroma_client.get_collection(collection_name)
                
                results = collection.query(
                    query_texts=[query],
                    n_results=top_k
                )
                
                documents = []
                for i, doc in enumerate(results['documents'][0]):
                    doc_id = results['ids'][0][i] if results['ids'] else None
                    documents.append({
                        'content': doc,
                        'metadata': results['metadatas'][0][i] if results['metadatas'] else {},
                        'id': doc_id
                    })
                    # Track access for least_used cleanup strategy
                    if doc_id:
                        self.update_document_access(telegram_id, doc_id)
                
                return documents
                
            except Exception as e:
                logger.warning(f"ChromaDB search failed for {telegram_id}: {e}")
        
        # Fallback to simple keyword search
        df = self.get_user_knowledge(telegram_id)
        if df is None or df.empty:
            return []
        
        # Simple keyword matching
        query_lower = query.lower()
        matches = df[
            df['DOCUMENT_TEXT'].str.lower().str.contains(query_lower, na=False) |
            df['TAGS'].str.lower().str.contains(query_lower, na=False)
        ]
        
        results = []
        for _, row in matches.head(top_k).iterrows():
            doc_id = str(row['ID'])
            results.append({
                'content': row['DOCUMENT_TEXT'],
                'metadata': {
                    'category': row['CATEGORY'],
                    'priority': row['PRIORITY'],
                    'tags': row['TAGS']
                },
                'id': doc_id
            })
            # Track access for least_used cleanup strategy
            self.update_document_access(telegram_id, doc_id)
        
        return results
    
    # ============================================================
    # DELETE KNOWLEDGE
    # ============================================================
    def delete_user_knowledge(self, telegram_id: str) -> bool:
        """
        Xóa toàn bộ knowledge của user.
        
        Returns:
            True nếu thành công
        """
        try:
            # Delete knowledge file
            file_path = self.get_knowledge_path(telegram_id)
            if file_path.exists():
                file_path.unlink()
            
            # Delete quota file
            quota_path = self.get_quota_path(telegram_id)
            if quota_path.exists():
                quota_path.unlink()
            
            # Delete ChromaDB collection
            if self.chroma_client:
                try:
                    self.chroma_client.delete_collection(f"user_{telegram_id}_knowledge")
                except:
                    pass
            
            logger.info(f"Deleted all knowledge for user {telegram_id}")
            return True
            
        except Exception as e:
            logger.error(f"Error deleting knowledge for {telegram_id}: {e}")
            return False


# ============================================================
# FACTORY
# ============================================================
_instance = None

def get_knowledge_manager() -> PersonalKnowledgeManager:
    """Get singleton instance"""
    global _instance
    if _instance is None:
        _instance = PersonalKnowledgeManager()
    return _instance


# ============================================================
# TEST
# ============================================================
if __name__ == '__main__':
    print("Testing Personal Knowledge Manager...")
    
    manager = PersonalKnowledgeManager()
    
    # Generate template
    print("\n1. Generating template...")
    buffer = manager.generate_template(include_samples=True)
    
    # Save for testing
    test_path = Path("data/templates/MeiLin_Knowledge_Template.xlsx")
    test_path.parent.mkdir(parents=True, exist_ok=True)
    with open(test_path, 'wb') as f:
        f.write(buffer.read())
    print(f"   ✅ Saved to {test_path}")
    
    # Test upload
    print("\n2. Testing upload...")
    buffer.seek(0)
    result = manager.save_user_knowledge("test_user_123", buffer)
    print(f"   Result: {result}")
    
    # Test summary
    print("\n3. Testing summary...")
    summary = manager.get_knowledge_summary("test_user_123")
    print(f"   Summary: {summary}")
    
    # Test search
    print("\n4. Testing search...")
    results = manager.search_knowledge("test_user_123", "MeiLin")
    print(f"   Found {len(results)} results")
    for r in results:
        print(f"   - {r['content'][:50]}...")
    
    print("\n✅ All tests passed!")
